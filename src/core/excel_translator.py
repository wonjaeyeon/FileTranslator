#!/usr/bin/env python3
"""
Excel 파일 번역기 - 리팩토링된 버전
서식 100% 보존하면서 내용만 번역
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os
import shutil
import tempfile
from typing import Optional, Callable, Set
from .translator import TranslationEngine

class ExcelTranslator:
    def __init__(self, progress_callback: Optional[Callable] = None):
        self.progress_callback = progress_callback or (lambda msg, pct: None)
        self.translator = TranslationEngine(progress_callback)
    
    def translate_file(self, 
                      input_path: str, 
                      output_path: str, 
                      direction: str,
                      preserve_english: bool = True,
                      add_new_sheet: bool = False,
                      exclude_cells: Optional[Set[str]] = None,
                      exclude_sheets: Optional[Set[str]] = None,
                      exclude_patterns: Optional[Set[str]] = None) -> bool:
        """
        Excel 파일 번역
        
        Args:
            input_path: 입력 파일 경로
            output_path: 출력 파일 경로  
            direction: 번역 방향 (ko_to_zh, zh_to_ko)
            preserve_english: 영문 보존 여부
            add_new_sheet: 새 시트 추가 여부
            exclude_cells: 제외할 셀 목록
            exclude_sheets: 제외할 시트 목록
            exclude_patterns: 제외할 패턴 목록
        """
        try:
            self.progress_callback("Excel 파일 로드 중...", 5)
            
            # 원본 파일 복사 (서식 보존)
            shutil.copy2(input_path, output_path)
            
            # 번역된 파일 작업
            workbook = openpyxl.load_workbook(output_path)
            sheet_names = workbook.sheetnames
            total_sheets = len(sheet_names)
            
            self.progress_callback(f"총 {total_sheets}개 시트 번역 시작", 10)
            
            for sheet_idx, sheet_name in enumerate(sheet_names):
                # 시트 제외 확인
                if exclude_sheets and sheet_name in exclude_sheets:
                    self.progress_callback(f"시트 '{sheet_name}' 제외됨", 
                                         10 + (sheet_idx / total_sheets) * 80)
                    continue
                
                sheet = workbook[sheet_name]
                original_sheet_name = sheet_name
                
                # 새 시트 추가 모드
                if add_new_sheet:
                    if direction == "ko_to_zh":
                        new_sheet_name = f"{sheet_name}_中文"
                    else:
                        new_sheet_name = f"{sheet_name}_한글"
                    
                    # 시트 복사
                    new_sheet = workbook.copy_worksheet(sheet)
                    new_sheet.title = new_sheet_name
                    sheet = new_sheet
                
                # 시트 내용 번역
                self._translate_sheet_content(
                    sheet, direction, preserve_english, 
                    sheet_idx, total_sheets, exclude_cells, 
                    exclude_patterns, original_sheet_name
                )
            
            # 파일 저장
            self.progress_callback("번역 완료, 파일 저장 중...", 95)
            workbook.save(output_path)
            workbook.close()
            
            self.progress_callback("번역이 완료되었습니다!", 100)
            return True
            
        except Exception as e:
            self.progress_callback(f"번역 중 오류 발생: {str(e)}", 0)
            return False
    
    def _translate_sheet_content(self, 
                               sheet, 
                               direction: str, 
                               preserve_english: bool,
                               sheet_idx: int, 
                               total_sheets: int,
                               exclude_cells: Optional[Set[str]] = None,
                               exclude_patterns: Optional[Set[str]] = None,
                               original_sheet_name: Optional[str] = None):
        """시트 내용 번역"""
        
        sheet_name_for_exclusion = original_sheet_name or sheet.title
        exclude_cells = exclude_cells or set()
        exclude_patterns = exclude_patterns or set()
        
        # 사용된 범위 계산
        if sheet.max_row == 1 and sheet.max_column == 1:
            return
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        total_cells = max_row * max_col
        
        if total_cells == 0:
            return
        
        processed_cells = 0
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                
                if cell.value is None:
                    processed_cells += 1
                    continue
                
                # 제외 확인
                cell_ref = f"{sheet_name_for_exclusion}!{get_column_letter(col)}{row}"
                if cell_ref in exclude_cells:
                    processed_cells += 1
                    continue
                
                # 패턴 제외 확인
                if self._should_exclude_by_pattern(str(cell.value), exclude_patterns):
                    processed_cells += 1
                    continue
                
                # 텍스트 번역
                if isinstance(cell.value, str) and cell.value.strip():
                    translated = self.translator.translate_text(
                        cell.value, direction, preserve_english
                    )
                    if translated != cell.value:
                        cell.value = translated
                
                processed_cells += 1
                
                # 진행률 업데이트
                sheet_progress = (processed_cells / total_cells) * 100
                overall_progress = 10 + (sheet_idx / total_sheets) * 80 + (sheet_progress / total_sheets)
                self.progress_callback(
                    f"시트 '{sheet.title}' 번역 중... ({processed_cells}/{total_cells})",
                    min(overall_progress, 90)
                )
    
    def _should_exclude_by_pattern(self, text: str, exclude_patterns: Set[str]) -> bool:
        """패턴에 의한 제외 여부 확인"""
        if not exclude_patterns:
            return False
        
        for pattern in exclude_patterns:
            if pattern.lower() in text.lower():
                return True
        return False
#!/usr/bin/env python3
"""
엑셀 파일 번역기 - 서식 완전 보존
openpyxl을 사용하여 원본 엑셀 파일의 모든 서식을 유지하면서 번역
"""

import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import requests
import re
import os
import time

class ExcelTranslator:
    def __init__(self):
        self.libretranslate_urls = [
            "https://libretranslate.de/translate",
            "https://translate.argosopentech.com/translate",
            "https://libretranslate.com/translate"
        ]

class ExcelTranslatorWithProgress(ExcelTranslator):
    def __init__(self, progress_callback=None):
        super().__init__()
        self.progress_callback = progress_callback or (lambda msg, pct: None)
    
    def translate_excel_file(self, input_path, output_path, direction='ko-zh', preserve_english=True, add_new_sheet=True):
        """진행률 콜백이 포함된 엑셀 파일 번역"""
        self.progress_callback("파일 열기 중...", 0)
        print(f"파일 열기: {input_path}")
        workbook = openpyxl.load_workbook(input_path)
        
        total_sheets = len(workbook.sheetnames)
        
        if add_new_sheet:
            # 새 시트에 번역본 추가
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                source_ws = workbook[sheet_name]
                translated_sheet_name = f"{sheet_name}_中文" if direction == 'ko-zh' else f"{sheet_name}_한국어"
                
                sheet_progress = (sheet_idx / total_sheets) * 90  # 90%까지 시트 처리
                self.progress_callback(f"시트 '{sheet_name}' 번역 중...", sheet_progress)
                print(f"시트 '{sheet_name}' 번역 중...")
                
                # 새 시트 생성
                target_ws = workbook.create_sheet(title=translated_sheet_name)
                
                # 완전한 시트 복사
                self.copy_worksheet(source_ws, target_ws)
                
                # 텍스트만 번역하여 업데이트
                self.translate_worksheet_with_progress(target_ws, direction, preserve_english, sheet_idx, total_sheets)
        else:
            # 원본 시트를 번역본으로 교체
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                source_ws = workbook[sheet_name]
                
                sheet_progress = (sheet_idx / total_sheets) * 90
                self.progress_callback(f"시트 '{sheet_name}' 번역 중...", sheet_progress)
                print(f"시트 '{sheet_name}' 번역 중...")
                
                self.translate_worksheet_with_progress(source_ws, direction, preserve_english, sheet_idx, total_sheets)

        self.progress_callback("파일 저장 중...", 95)
        print(f"번역 완료. 파일 저장: {output_path}")
        workbook.save(output_path)
        workbook.close()
        
        self.progress_callback("번역 완료!", 100)
        return output_path

    def translate_worksheet_with_progress(self, worksheet, direction, preserve_english, sheet_idx, total_sheets):
        """진행률 표시와 함께 워크시트 번역"""
        total_cells = worksheet.max_row * worksheet.max_column
        current_cell = 0
        
        for row in worksheet.iter_rows():
            for cell in row:
                current_cell += 1
                
                # 50셀마다 진행률 업데이트
                if current_cell % 50 == 0 or current_cell == total_cells:
                    cell_progress = (current_cell / total_cells) * 80 / total_sheets  # 시트당 80%/총시트수
                    total_progress = (sheet_idx / total_sheets) * 80 + cell_progress
                    self.progress_callback(f"번역 중... ({current_cell}/{total_cells} 셀)", total_progress)
                
                if cell.value and isinstance(cell.value, str):
                    translated = self.translate_text(cell.value, direction, preserve_english)
                    cell.value = translated

        # 발주서 전용 번역 사전 (부모 클래스에서 상속)
        
class ExcelTranslator:
    def __init__(self):
        self.libretranslate_urls = [
            "https://libretranslate.de/translate",
            "https://translate.argosopentech.com/translate",
            "https://libretranslate.com/translate"
        ]
        
        # 발주서 전용 번역 사전
        self.ko_to_zh_dict = {
            '발주서': '订单书',
            '수주처': '接单处',
            '상호': '商号',
            '대표': '代表',
            '발주일': '订单日期',
            '이메일': '邮箱',
            '연락처': '联系方式',
            '주소': '地址',
            '납기일자': '交货日期',
            '발송정보': '配送信息',
            '발송일': '发货日',
            '품목': '品目',
            '단위': '单位',
            '수량': '数量',
            '구분': '区分',
            '비고': '备注',
            '합계': '合计',
            '요구사항': '要求事项',
            '확인': '确认',
            '아래와 같이 발주합니다': '订单如下',
            '주식회사': '股份有限公司',
            '테클라스트코리아': '泰克拉斯特韩国',
            '이상모': '李相模',
            '등록번호': '注册号码',
            '경기도': '京畿道',
            '광명시': '光明市',
            '하안로': '下安路',
            '광명테크노파크': '光明科技园',
            '서비스': '服务',
            '도소매': '批发零售',
            '종목': '种目',
            '태블릿PC': '平板电脑',
            '유재건부장': '刘在建部长',
            '심대용과장': '沈大龙科长',
            '반입분': '入库分',
            '남품장소': '南品场所',
            '남품일정': '南品日程'
        }
        
        # 중국어 → 한국어 사전
        self.zh_to_ko_dict = {v: k for k, v in self.ko_to_zh_dict.items()}

    def is_korean(self, text):
        """한국어 포함 여부 검사"""
        return bool(re.search('[가-힣]', text))

    def is_chinese(self, text):
        """중국어 포함 여부 검사"""
        return bool(re.search('[\u4e00-\u9fff]', text))

    def is_english(self, text):
        """영문 여부 검사"""
        return bool(re.match(r'^[a-zA-Z0-9\s\.\,\-\(\)\[\]\{\}@:\/]+$', text.strip()))

    def translate_with_api(self, text, source_lang, target_lang):
        """LibreTranslate API로 번역"""
        for url in self.libretranslate_urls:
            try:
                data = {
                    "q": text,
                    "source": source_lang,
                    "target": target_lang,
                    "format": "text"
                }
                
                response = requests.post(url, json=data, timeout=10)
                if response.status_code == 200:
                    result = response.json()
                    if 'translatedText' in result:
                        return result['translatedText']
            except Exception as e:
                print(f"API 오류 ({url}): {e}")
                continue
        return None

    def fallback_translate(self, text, direction):
        """백업 번역 사전 사용"""
        if direction == 'ko-zh':
            dictionary = self.ko_to_zh_dict
        else:
            dictionary = self.zh_to_ko_dict
        
        translated = text
        for original, translation in dictionary.items():
            translated = translated.replace(original, translation)
        return translated

    def translate_text(self, text, direction, preserve_english=True):
        """텍스트 번역"""
        if not text or not isinstance(text, str) or not text.strip():
            return text
        
        # 영문 유지 옵션
        if preserve_english and self.is_english(text):
            return text
        
        # 언어 검사
        if direction == 'ko-zh' and not self.is_korean(text):
            return text
        elif direction == 'zh-ko' and not self.is_chinese(text):
            return text
        
        # API 번역 시도
        source_lang = 'ko' if direction == 'ko-zh' else 'zh'
        target_lang = 'zh' if direction == 'ko-zh' else 'ko'
        
        api_result = self.translate_with_api(text, source_lang, target_lang)
        if api_result:
            return api_result
        
        # 백업 사전 사용
        return self.fallback_translate(text, direction)

    def copy_worksheet(self, source_ws, target_ws):
        """워크시트 완전 복사 (모든 서식 포함)"""
        # 셀 데이터와 서식 복사
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = target_ws.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                
                # 서식 복사
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # 병합된 셀 복사
        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))

        # 행/열 크기 복사
        for row_num in range(1, source_ws.max_row + 1):
            if source_ws.row_dimensions[row_num].height:
                target_ws.row_dimensions[row_num].height = source_ws.row_dimensions[row_num].height

        for col_num in range(1, source_ws.max_column + 1):
            col_letter = get_column_letter(col_num)
            if source_ws.column_dimensions[col_letter].width:
                target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width

        # 시트 설정 복사
        target_ws.sheet_format = copy(source_ws.sheet_format)
        target_ws.sheet_properties = copy(source_ws.sheet_properties)

    def translate_excel_file(self, input_path, output_path, direction='ko-zh', preserve_english=True, add_new_sheet=True):
        """엑셀 파일 번역"""
        print(f"파일 열기: {input_path}")
        workbook = openpyxl.load_workbook(input_path)
        
        if add_new_sheet:
            # 새 시트에 번역본 추가
            for sheet_name in workbook.sheetnames:
                source_ws = workbook[sheet_name]
                translated_sheet_name = f"{sheet_name}_中文" if direction == 'ko-zh' else f"{sheet_name}_한국어"
                
                print(f"시트 '{sheet_name}' 번역 중...")
                
                # 새 시트 생성
                target_ws = workbook.create_sheet(title=translated_sheet_name)
                
                # 완전한 시트 복사
                self.copy_worksheet(source_ws, target_ws)
                
                # 텍스트만 번역하여 업데이트
                total_cells = target_ws.max_row * target_ws.max_column
                current_cell = 0
                
                for row in target_ws.iter_rows():
                    for cell in row:
                        current_cell += 1
                        if current_cell % 50 == 0:
                            print(f"진행률: {current_cell}/{total_cells} ({current_cell/total_cells*100:.1f}%)")
                        
                        if cell.value and isinstance(cell.value, str):
                            translated = self.translate_text(cell.value, direction, preserve_english)
                            cell.value = translated
        else:
            # 원본 시트를 번역본으로 교체
            for sheet_name in workbook.sheetnames:
                source_ws = workbook[sheet_name]
                
                print(f"시트 '{sheet_name}' 번역 중...")
                
                total_cells = source_ws.max_row * source_ws.max_column
                current_cell = 0
                
                for row in source_ws.iter_rows():
                    for cell in row:
                        current_cell += 1
                        if current_cell % 50 == 0:
                            print(f"진행률: {current_cell}/{total_cells} ({current_cell/total_cells*100:.1f}%)")
                        
                        if cell.value and isinstance(cell.value, str):
                            translated = self.translate_text(cell.value, direction, preserve_english)
                            cell.value = translated

        print(f"번역 완료. 파일 저장: {output_path}")
        workbook.save(output_path)
        workbook.close()
        return output_path

if __name__ == "__main__":
    translator = ExcelTranslator()
    
    # 테스트
    input_file = "sample-purchase-order.xlsx"
    output_file = "sample-purchase-order_translated.xlsx"
    
    if os.path.exists(input_file):
        result = translator.translate_excel_file(
            input_path=input_file,
            output_path=output_file,
            direction='ko-zh',
            preserve_english=True,
            add_new_sheet=True
        )
        print(f"번역 완료: {result}")
    else:
        print(f"파일을 찾을 수 없습니다: {input_file}")
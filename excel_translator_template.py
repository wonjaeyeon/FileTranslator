#!/usr/bin/env python3
"""
엑셀 파일 번역기 - 템플릿 방식
원본 파일을 직접 복사한 후 내용만 교체하여 서식 100% 보존
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os
import re
import shutil
import tempfile

class ExcelTranslatorTemplate:
    def __init__(self, progress_callback=None):
        self.progress_callback = progress_callback or (lambda msg, pct: None)
        
        # 확장된 한국어-중국어 번역 사전
        self.ko_to_zh_dict = {
            # 기본 발주서 용어
            '발주서': '订单书',
            '수주처': '接单处',
            '상호': '商号',
            '대표': '代表',
            '발주일': '订单日期',
            '이메일': '邮箱',
            '연락처': '联系方式',
            '주소': '地址',
            '납기일자': '交货日期',
            '발송정보': '配送信息',
            '발송일': '发货日',
            '품목': '品目',
            '단위': '单位',
            '수량': '数量',
            '구분': '区分',
            '비고': '备注',
            '합계': '合计',
            '요구사항': '要求事项',
            '확인': '确认',
            '아래와 같이 발주합니다': '订单如下',
            
            # 회사 정보
            '주식회사': '股份有限公司',
            '테클라스트코리아': '泰克拉스特韩国',
            '이상모': '李相模',
            '등록번호': '注册号码',
            '경기도': '京畿道',
            '광명시': '光明市',
            '하안로': '下安路',
            '광명테크노파크': '光明科技园',
            '서비스': '服务',
            '도소매': '批发零售',
            '종목': '种目',
            '태블릿PC': '平板电脑',
            '유재건부장': '刘在建部长',
            '심대용과장': '沈大龙科长',
            
            # 추가 업무 용어
            '반입분': '入库分',
            '남품장소': '南品场所',
            '남품일정': '南品日程',
            '발송정보': '配送信息',
            '반입분': '入库分',
            '남품장소': '南品场所',
            '남품일정': '南品日程',
            '전화번호': '电话号码',
            
            # 일반적인 한국어 단어들
            '회사': '公司',
            '기업': '企业',
            '제품': '产品',
            '상품': '商品',
            '가격': '价格',
            '금액': '金额',
            '비용': '费用',
            '날짜': '日期',
            '시간': '时间',
            '장소': '地点',
            '위치': '位置',
            '담당자': '负责人',
            '관리자': '管理者',
            '직원': '职员',
            '부장': '部长',
            '과장': '科长',
            '대리': '代理',
            '사원': '职员',
            '팀장': '组长',
            '부서': '部门',
            '팀': '团队',
            '업무': '业务',
            '작업': '工作',
            '계획': '计划',
            '예정': '预定',
            '완료': '完成',
            '진행': '进行',
            '시작': '开始',
            '종료': '结束',
            '처리': '处理',
            '검토': '审查',
            '승인': '批准',
            '취소': '取消',
            '수정': '修改',
            '변경': '变更',
            '추가': '添加',
            '삭제': '删除',
            '입력': '输入',
            '출력': '输出',
            '전송': '传送',
            '수신': '接收',
            '발신': '发送',
            '문서': '文件',
            '서류': '资料',
            '양식': '格式',
            '형식': '格式',
            '내용': '内容',
            '정보': '信息',
            '데이터': '数据',
            '자료': '资料',
            '참고': '参考',
            '첨부': '附件',
            '링크': '链接',
            '연결': '连接',
            '접속': '连接',
            '로그인': '登录',
            '로그아웃': '退出',
            '계정': '账户',
            '사용자': '用户',
            '고객': '客户',
            '구매자': '购买者',
            '판매자': '销售者',
            '공급자': '供应商',
            '제조사': '制造商',
            '생산자': '生产者',
            '배송': '配送',
            '운송': '运输',
            '물류': '物流',
            '창고': '仓库',
            '재고': '库存',
            '입고': '入库',
            '출고': '出库',
            '검사': '检查',
            '품질': '质量',
            '검증': '验证',
            '테스트': '测试',
            '시험': '试验',
            '결과': '结果',
            '성과': '成果',
            '실적': '业绩',
            '성공': '成功',
            '실패': '失败',
            '오류': '错误',
            '문제': '问题',
            '해결': '解决',
            '개선': '改善',
            '향상': '提升',
            '발전': '发展',
            '성장': '成长',
            '확대': '扩大',
            '축소': '缩小',
            '증가': '增加',
            '감소': '减少',
            '상승': '上升',
            '하락': '下降',
            '안정': '稳定',
            '변동': '变动',
            '조정': '调整',
            '설정': '设定',
            '구성': '构成',
            '설치': '安装',
            '제거': '删除',
            '업데이트': '更新',
            '업그레이드': '升级',
            '다운로드': '下载',
            '업로드': '上传',
            '저장': '保存',
            '백업': '备份',
            '복원': '恢复',
            '복사': '复制',
            '이동': '移动',
            '붙여넣기': '粘贴',
            '잘라내기': '剪切',
            '선택': '选择',
            '취소선택': '取消选择',
            '전체선택': '全部选择',
            '검색': '搜索',
            '찾기': '查找',
            '필터': '筛选',
            '정렬': '排序',
            '분류': '分类',
            '그룹': '组',
            '카테고리': '类别',
            '항목': '项目',
            '목록': '列表',
            '리스트': '列表',
            '메뉴': '菜单',
            '옵션': '选项',
            '설정값': '设定值',
            '기본값': '默认值',
            '최대값': '最大值',
            '최소값': '最小值',
            '평균값': '平均值',
            '합계값': '合计值',
            '총계': '总计',
            '소계': '小计',
            '세부내역': '详细内容',
            '상세정보': '详细信息',
            '요약': '摘要',
            '개요': '概要',
            '소개': '介绍',
            '설명': '说明',
            '안내': '指导',
            '지침': '指南',
            '규칙': '规则',
            '규정': '规定',
            '정책': '政策',
            '방침': '方针',
            '원칙': '原则',
            '기준': '标准',
            '조건': '条件',
            '요구사항': '要求事项',
            '필수사항': '必需事项',
            '선택사항': '选择事项',
            '권장사항': '推荐事项',
            '주의사항': '注意事项',
            '경고': '警告',
            '알림': '通知',
            '공지': '通知',
            '발표': '发表',
            '보고': '报告',
            '발표자료': '发表资料',
            '보고서': '报告书',
            '제안서': '提案书',
            '계획서': '计划书',
            '명세서': '明细书',
            '사양서': '规格书',
            '지시서': '指示书',
            '안내서': '指南书'
        }
        
        # 중국어 → 한국어 사전
        self.zh_to_ko_dict = {v: k for k, v in self.ko_to_zh_dict.items()}

    def is_korean(self, text):
        return bool(re.search('[가-힣]', str(text)))

    def is_chinese(self, text):
        return bool(re.search('[\u4e00-\u9fff]', str(text)))

    def is_english(self, text):
        return bool(re.match(r'^[a-zA-Z0-9\s\.\,\-\(\)\[\]\{\}@:\/]+$', str(text).strip()))
    
    def is_cell_excluded(self, cell, exclude_cells, sheet_name):
        """셀이 제외 범위에 포함되는지 확인"""
        if not exclude_cells:
            return False
            
        cell_addr = cell.coordinate
        
        for exclude_range in exclude_cells:
            exclude_range = exclude_range.strip()
            if not exclude_range:
                continue
                
            # 시트명이 포함된 경우 (Sheet1!A1 형식)
            if '!' in exclude_range:
                sheet_part, cell_part = exclude_range.split('!', 1)
                if sheet_part != sheet_name:
                    continue  # 다른 시트의 셀이면 건너뜀
                exclude_range = cell_part
            
            if ':' in exclude_range:
                # 범위 형식 (예: A1:B5)
                try:
                    from openpyxl.utils import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(exclude_range)
                    if min_col <= cell.column <= max_col and min_row <= cell.row <= max_row:
                        return True
                except:
                    pass
            elif exclude_range == cell_addr:
                # 단일 셀 (예: A1)
                return True
        return False
    
    def is_text_excluded(self, text, exclude_patterns):
        """텍스트가 제외 패턴에 포함되는지 확인"""
        text_lower = str(text).lower()
        for pattern in exclude_patterns:
            pattern = pattern.strip().lower()
            if pattern in text_lower:
                return True
        return False

    def translate_with_libretranslate(self, text, source_lang, target_lang):
        """LibreTranslate API로 번역"""
        libretranslate_urls = [
            "https://libretranslate.de/translate",
            "https://translate.argosopentech.com/translate", 
            "https://libretranslate.com/translate"
        ]
        
        for url in libretranslate_urls:
            try:
                import requests
                data = {
                    "q": text,
                    "source": source_lang,
                    "target": target_lang,
                    "format": "text"
                }
                
                response = requests.post(url, json=data, timeout=8)
                if response.status_code == 200:
                    result = response.json()
                    if 'translatedText' in result and result['translatedText'].strip():
                        return result['translatedText']
            except Exception:
                continue
        return None

    def translate_with_google(self, text, direction):
        """Google Translate로 번역"""
        try:
            from googletrans import Translator
            translator = Translator()
            
            if direction == 'ko-zh':
                source, target = 'ko', 'zh-CN'
            else:
                source, target = 'zh-CN', 'ko'
            
            result = translator.translate(text, src=source, dest=target)
            if result and result.text and result.text.strip() != text:
                return result.text.strip()
        except Exception:
            pass
        return None

    def translate_with_huggingface(self, text, direction):
        """Hugging Face 무료 API로 번역"""
        try:
            import requests
            
            # Hugging Face의 무료 번역 모델들
            if direction == 'ko-zh':
                model_id = "Helsinki-NLP/opus-mt-ko-zh"
            else:
                model_id = "Helsinki-NLP/opus-mt-zh-ko"
            
            api_url = f"https://api-inference.huggingface.co/models/{model_id}"
            
            response = requests.post(
                api_url,
                headers={},  # API 키 없이도 사용 가능 (제한적)
                json={"inputs": text},
                timeout=10
            )
            
            if response.status_code == 200:
                result = response.json()
                if isinstance(result, list) and len(result) > 0:
                    translation = result[0].get('translation_text', '')
                    if translation.strip():
                        return translation
        except Exception:
            pass
        return None

    def translate_with_ollama(self, text, direction):
        """Ollama 로컬 LLM으로 번역"""
        try:
            import requests
            
            # Ollama가 로컬에서 실행 중인지 확인
            target_lang = "중국어" if direction == 'ko-zh' else "한국어"
            source_lang = "한국어" if direction == 'ko-zh' else "중국어"
            
            prompt = f"다음 {source_lang} 텍스트를 {target_lang}로 번역해주세요. 번역 결과만 출력하세요:\n\n{text}"
            
            response = requests.post(
                "http://localhost:11434/api/generate",
                json={
                    "model": "llama3.2:3b",  # 가벼운 모델
                    "prompt": prompt,
                    "stream": False
                },
                timeout=15
            )
            
            if response.status_code == 200:
                result = response.json()
                translation = result.get('response', '').strip()
                if translation and translation != text:
                    return translation
        except Exception:
            pass
        return None

    def translate_text(self, text, direction, preserve_english=True):
        if not text or not isinstance(text, str) or not str(text).strip():
            return text
        
        text_str = str(text).strip()
        
        # 영문만으로 구성된 텍스트는 유지
        if preserve_english and self.is_english(text_str):
            return text
        
        # 혼합 텍스트 처리 (영문 + 한국어/중국어)
        if direction == 'ko-zh':
            if not self.is_korean(text_str):
                return text  # 한국어가 없으면 번역하지 않음
        elif direction == 'zh-ko':
            if not self.is_chinese(text_str):
                return text  # 중국어가 없으면 번역하지 않음
        
        # 1단계: 번역 사전 먼저 시도 (빠르고 정확)
        dictionary = self.ko_to_zh_dict if direction == 'ko-zh' else self.zh_to_ko_dict
        sorted_keys = sorted(dictionary.keys(), key=len, reverse=True)
        
        translated = text_str
        has_dict_translation = False
        
        for original in sorted_keys:
            if original in translated:
                translated = translated.replace(original, dictionary[original])
                has_dict_translation = True
        
        # 사전 번역 후에도 원본 언어가 남아있는지 확인
        still_has_original_lang = False
        if direction == 'ko-zh' and self.is_korean(translated):
            still_has_original_lang = True
        elif direction == 'zh-ko' and self.is_chinese(translated):
            still_has_original_lang = True
        
        # 사전만으로 완전 번역되면 사전 결과 반환
        if has_dict_translation and not still_has_original_lang:
            return translated
        
        # 2단계: Google Translate 시도 (가장 안정적)
        google_translation = self.translate_with_google(text_str, direction)
        if google_translation:
            return google_translation
        
        # 3단계: LibreTranslate API 시도
        source_lang = 'ko' if direction == 'ko-zh' else 'zh'
        target_lang = 'zh' if direction == 'ko-zh' else 'ko'
        
        api_translation = self.translate_with_libretranslate(text_str, source_lang, target_lang)
        if api_translation:
            return api_translation
        
        # 4단계: Hugging Face API 시도
        hf_translation = self.translate_with_huggingface(text_str, direction)
        if hf_translation:
            return hf_translation
        
        # 5단계: Ollama 로컬 LLM 시도
        ollama_translation = self.translate_with_ollama(text_str, direction)
        if ollama_translation:
            return ollama_translation
        
        # 6단계: 모든 방법 실패시 사전 번역 결과라도 반환
        return translated

    def translate_excel_file(self, input_path, output_path, direction='ko-zh', preserve_english=True, add_new_sheet=True, exclude_sheets=None, exclude_cells=None, exclude_patterns=None):
        """템플릿 방식 엑셀 번역 - 원본 파일 복사 후 내용만 교체"""
        
        self.progress_callback("원본 파일 복사 중...", 0)
        
        # 1단계: 원본 파일을 출력 파일로 직접 복사
        shutil.copy2(input_path, output_path)
        
        self.progress_callback("번역 작업 준비 중...", 10)
        
        # 2단계: 복사된 파일에서 텍스트만 번역하여 교체
        # openpyxl로 열어서 텍스트만 교체 (서식은 건드리지 않음)
        workbook = openpyxl.load_workbook(output_path, data_only=False)
        
        total_sheets = len(workbook.sheetnames)
        
        if add_new_sheet:
            # 새 시트 추가 방식
            original_sheets = list(workbook.sheetnames)
            
            for sheet_idx, sheet_name in enumerate(original_sheets):
                # 제외할 시트인지 확인
                if exclude_sheets and sheet_name in exclude_sheets:
                    self.progress_callback(f"시트 '{sheet_name}' 건너뜀 (제외 목록)", (sheet_idx / len(original_sheets)) * 40 + 20)
                    continue
                    
                sheet = workbook[sheet_name]
                sheet_progress = (sheet_idx / len(original_sheets)) * 40 + 20
                
                self.progress_callback(f"시트 '{sheet_name}' 복사 중...", sheet_progress)
                
                # 시트 복사
                new_sheet = workbook.copy_worksheet(sheet)
                translated_sheet_name = f"{sheet_name}_中文" if direction == 'ko-zh' else f"{sheet_name}_한국어"
                new_sheet.title = translated_sheet_name
                
                # 복사된 시트에서만 번역 (원본 시트명을 사용해서 제외 체크)
                self.translate_sheet_content_only(new_sheet, direction, preserve_english, sheet_idx, len(original_sheets), exclude_cells, exclude_patterns, original_sheet_name=sheet_name)
        else:
            # 원본 시트에서 직접 번역
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                # 제외할 시트인지 확인
                if exclude_sheets and sheet_name in exclude_sheets:
                    self.progress_callback(f"시트 '{sheet_name}' 건너뜀 (제외 목록)", (sheet_idx / total_sheets) * 60 + 20)
                    continue
                    
                sheet = workbook[sheet_name]
                self.translate_sheet_content_only(sheet, direction, preserve_english, sheet_idx, total_sheets, exclude_cells, exclude_patterns)
        
        self.progress_callback("변경사항 저장 중...", 90)
        
        # 3단계: 저장 (서식은 그대로, 텍스트만 변경됨)
        workbook.save(output_path)
        workbook.close()
        
        self.progress_callback("번역 완료!", 100)
        return output_path

    def translate_sheet_content_only(self, sheet, direction, preserve_english, sheet_idx, total_sheets, exclude_cells=None, exclude_patterns=None, original_sheet_name=None):
        """시트 내용만 번역 (서식은 건드리지 않음)"""
        
        # 원본 시트명 (새 시트 추가 모드에서 사용)
        sheet_name_for_exclusion = original_sheet_name or sheet.title
        
        print(f"시트 '{sheet.title}' 번역 시작 (제외 체크용 시트명: '{sheet_name_for_exclusion}')")
        if exclude_cells:
            print(f"  제외할 셀: {len(exclude_cells)}개 - {exclude_cells[:3]}{'...' if len(exclude_cells) > 3 else ''}")
        if exclude_patterns:
            print(f"  제외할 패턴: {exclude_patterns}")
        
        # 시트의 모든 셀을 순회하면서 값만 변경
        total_cells = 0
        cells_to_translate = []
        
        # 먼저 번역할 셀들 수집
        excluded_count = 0
        for row in sheet.iter_rows():
            for cell in row:
                total_cells += 1
                
                # 값이 있는 셀만 처리
                if cell.value and isinstance(cell.value, str) and cell.value.strip():
                    # 제외할 셀 범위 확인 (원본 시트명 사용)
                    if exclude_cells and self.is_cell_excluded(cell, exclude_cells, sheet_name_for_exclusion):
                        excluded_count += 1
                        print(f"  제외: {sheet_name_for_exclusion}!{cell.coordinate} = '{cell.value}'")
                        continue
                    
                    # 제외할 패턴 확인
                    if exclude_patterns and self.is_text_excluded(cell.value, exclude_patterns):
                        excluded_count += 1
                        print(f"  패턴 제외: {sheet_name_for_exclusion}!{cell.coordinate} = '{cell.value}'")
                        continue
                        
                    cells_to_translate.append(cell)
        
        print(f"  총 {total_cells}개 셀 중 {len(cells_to_translate)}개 번역 예정 ({excluded_count}개 제외)")
        
        # 번역 작업
        for idx, cell in enumerate(cells_to_translate):
            if idx % 10 == 0 or idx == len(cells_to_translate) - 1:
                sheet_base = (sheet_idx / total_sheets) * 60 + 20
                cell_progress = (idx / len(cells_to_translate)) * (60 / total_sheets) if cells_to_translate else 0
                total_progress = min(85, sheet_base + cell_progress)
                self.progress_callback(f"시트 번역 중... ({idx + 1}/{len(cells_to_translate)} 텍스트)", total_progress)
            
            original_value = cell.value
            translated_value = self.translate_text(original_value, direction, preserve_english)
            
            # 값이 실제로 변경된 경우에만 업데이트
            if translated_value != original_value:
                cell.value = translated_value

if __name__ == "__main__":
    def test_progress(msg, pct):
        print(f"[{pct:5.1f}%] {msg}")
    
    translator = ExcelTranslatorTemplate(test_progress)
    
    # 테스트
    input_file = "sample-purchase-order.xlsx"
    output_file = "sample-purchase-order_template.xlsx"
    
    if os.path.exists(input_file):
        try:
            result = translator.translate_excel_file(
                input_path=input_file,
                output_path=output_file,
                direction='ko-zh',
                preserve_english=True,
                add_new_sheet=True
            )
            print(f"번역 완료: {result}")
        except Exception as e:
            print(f"번역 오류: {e}")
    else:
        print(f"파일을 찾을 수 없습니다: {input_file}")
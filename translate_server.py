#!/usr/bin/env python3
"""
간단한 번역 프록시 서버
CORS 문제를 해결하고 안정적인 번역을 위해 사용
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import requests
import json
import time
import threading
import uuid

app = Flask(__name__)
CORS(app)

# 번역 작업 상태 저장
translation_jobs = {}

# LibreTranslate 공개 인스턴스들
LIBRETRANSLATE_URLS = [
    "https://libretranslate.de/translate",
    "https://translate.argosopentech.com/translate",
    "https://libretranslate.com/translate"
]

def translate_with_libretranslate(text, source_lang, target_lang):
    """LibreTranslate API를 사용한 번역"""
    for url in LIBRETRANSLATE_URLS:
        try:
            data = {
                "q": text,
                "source": source_lang,
                "target": target_lang,
                "format": "text"
            }
            
            response = requests.post(url, json=data, timeout=8)
            if response.status_code == 200:
                result = response.json()
                if 'translatedText' in result and result['translatedText'].strip():
                    return result['translatedText']
        except Exception:
            continue
    return None

def translate_with_google(text, direction):
    """Google Translate 무료 API로 번역"""
    try:
        from googletrans import Translator
        translator = Translator()
        
        if direction == 'ko-zh':
            source, target = 'ko', 'zh-CN'
        else:
            source, target = 'zh-CN', 'ko'
        
        result = translator.translate(text, src=source, dest=target)
        if result and result.text and result.text.strip() != text:
            return result.text.strip()
    except Exception:
        pass
    return None

def translate_with_deep_translator(text, direction):
    """Deep Translator로 번역 (Google, Bing, DeepL 등 지원)"""
    try:
        from deep_translator import GoogleTranslator, BingTranslator
        
        if direction == 'ko-zh':
            source, target = 'ko', 'zh-CN'
        else:
            source, target = 'zh-CN', 'ko'
        
        # Google Translator 시도
        try:
            translator = GoogleTranslator(source=source, target=target)
            result = translator.translate(text)
            if result and result.strip() != text:
                return result.strip()
        except:
            pass
        
        # Bing Translator 시도
        try:
            translator = BingTranslator(source=source, target=target)
            result = translator.translate(text)
            if result and result.strip() != text:
                return result.strip()
        except:
            pass
            
    except Exception:
        pass
    return None

def translate_with_ollama(text, direction):
    """Ollama 로컬 LLM으로 번역"""
    try:
        if direction == 'ko-zh':
            prompt = f"다음 한국어 텍스트를 중국어로 정확하게 번역해주세요. 영어는 그대로 유지하세요. 번역 결과만 답해주세요:\n\n{text}"
        else:
            prompt = f"다음 중국어 텍스트를 한국어로 정확하게 번역해주세요. 영어는 그대로 유지하세요. 번역 결과만 답해주세요:\n\n{text}"
        
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={
                "model": "llama3.2:3b",  # 가벼운 모델
                "prompt": prompt,
                "stream": False
            },
            timeout=15
        )
        
        if response.status_code == 200:
            result = response.json()
            translation = result.get('response', '').strip()
            if translation and translation != text:
                return translation
    except Exception:
        pass
    return None

def is_korean(text):
    import re
    return bool(re.search('[가-힣]', str(text)))

def is_chinese(text):
    import re
    return bool(re.search('[\u4e00-\u9fff]', str(text)))

def is_english(text):
    import re
    return bool(re.match(r'^[a-zA-Z0-9\s\.\,\-\(\)\[\]\{\}@:\/]+$', str(text).strip()))

def fallback_translation(text, direction):
    """백업 번역 사전"""
    if direction == 'ko-zh':
        translations = {
            '발주서': '订单书',
            '수주처': '接单处',
            '상호': '商号',
            '대표': '代表',
            '발주일': '订单日期',
            '이메일': '邮箱',
            '연락처': '联系方式',
            '주소': '地址',
            '납기일자': '交货日期',
            '발송정보': '配送信息',
            '발송일': '发货日',
            '품목': '品目',
            '단위': '单位',
            '수량': '数量',
            '구분': '区分',
            '비고': '备注',
            '합계': '合计',
            '요구사항': '要求事项',
            '확인': '确认',
            '아래와 같이 발주합니다': '订单如下',
            '주식회사': '股份有限公司',
            '테클라스트코리아': '泰克拉斯特韩国',
            '이상모': '李相模',
            '등록번호': '注册号码',
            '경기도': '京畿道',
            '광명시': '光明市',
            '하안로': '下安路',
            '광명테크노파크': '光明科技园',
            '서비스': '服务',
            '도소매': '批发零售',
            '종목': '种目',
            '태블릿PC': '平板电脑',
            '유재건부장': '刘在建部长',
            '심대용과장': '沈大龙科长',
            '반입분': '入库分',
            '남품장소': '南品场所',
            '남품일정': '南品日程',
            # 추가 번역 용어들
            '프로젝트': '项目',
            '업무': '业务',
            '관리자': '管理员',
            '완료': '完成',
            '고객': '客户',
            '회사': '公司',
            '부서': '部门',
            '담당자': '负责人',
            '직원': '职员',
            '팀장': '组长',
            '과장': '科长',
            '부장': '部长',
            '사장': '社长',
            '작업': '工作',
            '계획': '计划',
            '일정': '日程',
            '진행': '进行',
            '시작': '开始',
            '종료': '结束',
            '검토': '审查',
            '승인': '批准',
            '변경': '变更',
            '수정': '修改',
            '업데이트': '更新',
            '품질': '质量',
            '성능': '性能',
            '정보': '信息',
            '데이터': '数据',
            '문서': '文件',
            '보고서': '报告书',
            '제안서': '提案书',
            '계획서': '计划书',
            '내용': '内容',
            '항목': '项',
            '목록': '列表',
            '설정': '设置',
            '상태': '状态',
            '결과': '结果',
            '목적': '目的',
            '목표': '目标',
            '방법': '方法',
            '절차': '程序',
            '과정': '过程',
            '단계': '阶段',
            '관리': '管理',
            '점검': '检查',
            '테스트': '测试',
            '확인': '确认',
            '평가': '评价',
            '분석': '分析',
            '개발': '开发',
            '설계': '设计',
            '생산': '生产',
            '제조': '制造',
            '설치': '安装',
            '배송': '配送',
            '시스템': '系统',
            '장비': '设备',
            '제품': '产品',
            '모델': '型号',
            '종류': '种类',
            '크기': '尺寸',
            '시간': '时间',
            '기간': '期间',
            '가격': '价格',
            '비용': '费用',
            '금액': '金额',
            '총계': '总计',
            '기본': '基本',
            '표준': '标准',
            '특별': '特别',
            '현재': '现在',
            '새로운': '新的',
            '최신': '最新'
        }
    else:  # zh-ko
        translations = {
            '订单书': '발주서',
            '接单处': '수주처',
            '商号': '상호',
            '代表': '대표',
            '订单日期': '발주일',
            '邮箱': '이메일',
            '联系方式': '연락처',
            '地址': '주소',
            '交货日期': '납기일자',
            '配送信息': '발송정보',
            '发货日': '발송일',
            '品目': '품목',
            '单位': '단위',
            '数量': '수량',
            '区分': '구분',
            '备注': '비고',
            '合计': '합계',
            '要求事项': '요구사항',
            '确认': '확인',
            '订单如下': '아래와 같이 발주합니다',
            '股份有限公司': '주식회사',
            '泰克拉斯特韩国': '테클라스트코리아',
            '李相模': '이상모',
            '注册号码': '등록번호',
            '京畿道': '경기도',
            '光明市': '광명시',
            '下安路': '하안로',
            '光明科技园': '광명테크노파크',
            '服务': '서비스',
            '批发零售': '도소매',
            '种目': '종목',
            '平板电脑': '태블릿PC',
            '刘在建部长': '유재건부장',
            '沈大龙科长': '심대용과장',
            '入库分': '반입분',
            '南品场所': '남품장소',
            '南品日程': '남품일정',
            # 추가 번역 용어들 (중국어 -> 한국어)
            '项目': '프로젝트',
            '业务': '업무',
            '管理员': '관리자',
            '完成': '완료',
            '客户': '고객',
            '公司': '회사',
            '部门': '부서',
            '负责人': '담당자',
            '职员': '직원',
            '组长': '팀장',
            '科长': '과장',
            '部长': '부장',
            '社长': '사장',
            '工作': '작업',
            '计划': '계획',
            '日程': '일정',
            '进行': '진행',
            '开始': '시작',
            '结束': '종료',
            '审查': '검토',
            '批准': '승인',
            '变更': '변경',
            '修改': '수정',
            '更新': '업데이트',
            '质量': '품질',
            '性能': '성능',
            '信息': '정보',
            '数据': '데이터',
            '文件': '문서',
            '报告书': '보고서',
            '提案书': '제안서',
            '计划书': '계획서',
            '内容': '내용',
            '项': '항목',
            '列表': '목록',
            '设置': '설정',
            '状态': '상태',
            '结果': '결과',
            '目的': '목적',
            '目标': '목표',
            '方法': '방법',
            '程序': '절차',
            '过程': '과정',
            '阶段': '단계',
            '管理': '관리',
            '检查': '점검',
            '测试': '테스트',
            '确认': '확인',
            '评价': '평가',
            '分析': '분석',
            '开发': '개발',
            '设计': '설계',
            '生产': '생산',
            '制造': '제조',
            '安装': '설치',
            '配送': '배송',
            '系统': '시스템',
            '设备': '장비',
            '产品': '제품',
            '型号': '모델',
            '种类': '종류',
            '尺寸': '크기',
            '时间': '시간',
            '期间': '기간',
            '价格': '가격',
            '费用': '비용',
            '金额': '금액',
            '总计': '총계',
            '基本': '기본',
            '标准': '표준',
            '特别': '특별',
            '现在': '현재',
            '新的': '새로운',
            '最新': '최신'
        }
    
    translated = text
    for original, translation in translations.items():
        translated = translated.replace(original, translation)
    
    return translated

def hybrid_translate_text(text, direction, preserve_english=True):
    """하이브리드 번역 시스템 (사전 -> API -> LLM)"""
    if not text or not isinstance(text, str) or not str(text).strip():
        return text
    
    text_str = str(text).strip()
    
    # 영문만으로 구성된 텍스트는 유지
    if preserve_english and is_english(text_str):
        return text
    
    # 혼합 텍스트 처리 (영문 + 한국어/중국어)
    if direction == 'ko-zh':
        if not is_korean(text_str):
            return text  # 한국어가 없으면 번역하지 않음
    elif direction == 'zh-ko':
        if not is_chinese(text_str):
            return text  # 중국어가 없으면 번역하지 않음
    
    # 1단계: 번역 사전 먼저 시도 (빠르고 정확)
    dict_translated = fallback_translation(text_str, direction)
    has_dict_translation = dict_translated != text_str
    
    # 사전 번역 후에도 원본 언어가 남아있는지 확인
    still_has_original_lang = False
    if direction == 'ko-zh' and is_korean(dict_translated):
        still_has_original_lang = True
    elif direction == 'zh-ko' and is_chinese(dict_translated):
        still_has_original_lang = True
    
    # 사전만으로 완전 번역되면 사전 결과 반환
    if has_dict_translation and not still_has_original_lang:
        return dict_translated
    
    # 2단계: Google Translate 시도 (전체 문장 번역)
    google_translation = translate_with_google(text_str, direction)
    if google_translation:
        return google_translation
    
    # 3단계: Deep Translator (Google/Bing) 시도
    deep_translation = translate_with_deep_translator(text_str, direction)
    if deep_translation:
        return deep_translation
    
    # 4단계: LibreTranslate API 시도
    source_lang = 'ko' if direction == 'ko-zh' else 'zh'
    target_lang = 'zh' if direction == 'ko-zh' else 'ko'
    
    libre_translation = translate_with_libretranslate(text_str, source_lang, target_lang)
    if libre_translation:
        return libre_translation
    
    # 5단계: Ollama 로컬 LLM 시도
    ollama_translation = translate_with_ollama(text_str, direction)
    if ollama_translation:
        return ollama_translation
    
    # 6단계: 모든 방법 실패시 사전 번역 결과라도 반환
    return dict_translated

@app.route('/translate', methods=['POST'])
def translate():
    try:
        data = request.json
        text = data.get('text', '')
        direction = data.get('direction', 'ko-zh')
        preserve_english = data.get('preserve_english', True)
        
        if not text.strip():
            return jsonify({'translatedText': text})
        
        # 하이브리드 번역 시스템 사용
        translated = hybrid_translate_text(text, direction, preserve_english)
        
        # 어떤 방법이 사용되었는지 확인
        dict_only = fallback_translation(text, direction)
        if translated == text:
            method = 'unchanged'
        elif translated == dict_only:
            # 사전만으로 번역된 경우라도, 원본 언어가 남아있으면 API 필요
            if direction == 'ko-zh' and is_korean(translated):
                method = 'partial_dictionary'
            elif direction == 'zh-ko' and is_chinese(translated):
                method = 'partial_dictionary'
            else:
                method = 'dictionary'
        else:
            method = 'api_or_llm'
        
        return jsonify({
            'translatedText': translated, 
            'method': method,
            'original': text
        })
            
    except Exception as e:
        print(f"번역 오류: {e}")
        return jsonify({'error': str(e)}), 500

def run_translation(job_id, input_path, output_path, direction, preserve_english, add_new_sheet, exclude_sheets=None, exclude_cells=None, exclude_patterns=None):
    """백그라운드에서 번역 실행"""
    try:
        from excel_translator_template import ExcelTranslatorTemplate
        
        def progress_callback(message, percentage):
            translation_jobs[job_id]['progress'] = percentage
            translation_jobs[job_id]['message'] = message
            print(f"Job {job_id}: {message} ({percentage}%)")
        
        translation_jobs[job_id]['status'] = 'running'
        translation_jobs[job_id]['progress'] = 0
        translation_jobs[job_id]['message'] = '번역 시작'
        
        translator = ExcelTranslatorTemplate(progress_callback)
        result_path = translator.translate_excel_file(
            input_path=input_path,
            output_path=output_path,
            direction=direction,
            preserve_english=preserve_english,
            add_new_sheet=add_new_sheet,
            exclude_sheets=exclude_sheets,
            exclude_cells=exclude_cells,
            exclude_patterns=exclude_patterns
        )
        
        translation_jobs[job_id]['status'] = 'completed'
        translation_jobs[job_id]['progress'] = 100
        translation_jobs[job_id]['message'] = '번역 완료'
        translation_jobs[job_id]['result_path'] = result_path
        
        # 임시 파일 정리
        import os
        if os.path.exists(input_path):
            os.remove(input_path)
            
    except Exception as e:
        translation_jobs[job_id]['status'] = 'error'
        translation_jobs[job_id]['error'] = str(e)
        print(f"번역 오류 (Job {job_id}): {e}")

@app.route('/extract-words', methods=['POST'])
def extract_words():
    """엑셀 파일에서 번역 대상 단어들을 추출하여 <CELL, 단어> 형태로 반환"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '파일이 업로드되지 않았습니다.'}), 400

        file = request.files['file']
        direction = request.form.get('direction', 'ko-zh')

        if file.filename == '':
            return jsonify({'error': '파일이 선택되지 않았습니다.'}), 400

        # 파일 저장 - tmp 디렉토리 생성 및 원본 파일 완전 복사
        import os
        import openpyxl
        import shutil

        unique_id = str(uuid.uuid4())[:8]

        # tmp 디렉토리 생성
        tmp_dir = os.path.join(os.getcwd(), 'tmp')
        os.makedirs(tmp_dir, exist_ok=True)

        input_filename = f"input_{unique_id}_{file.filename}"
        input_path = os.path.join(tmp_dir, input_filename)

        file.save(input_path)

        # 파일 보존을 위해 복사본 생성 (모든 요소 보장)
        backup_filename = f"backup_{unique_id}_{file.filename}"
        backup_path = os.path.join(tmp_dir, backup_filename)
        shutil.copy2(input_path, backup_path)

        print(f"원본 파일 저장: {input_path}")
        print(f"백업 파일 생성: {backup_path}")

        # 엑셀 파일에서 단어 추출
        workbook = openpyxl.load_workbook(input_path)
        word_list = []

        print(f"엑셀 파일 시트 목록: {workbook.sheetnames}")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"시트 '{sheet_name}' 처리 중...")

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.strip():
                        cell_text = str(cell.value).strip()

                        # 번역이 필요한 텍스트인지 확인
                        needs_translation = False
                        if direction == 'ko-zh' and is_korean(cell_text):
                            needs_translation = True
                        elif direction == 'zh-ko' and is_chinese(cell_text):
                            needs_translation = True

                        if needs_translation:
                            cell_address = f"{sheet_name}!{cell.coordinate}"
                            word_list.append(f"<{cell_address}, {cell_text}>")
                            print(f"  추출: {cell_address} = {cell_text}")

        # 원본 파일은 제거하지 않고 보존 (백업 파일 유지)

        # GPT 프롬프트 생성
        target_lang = "중국어" if direction == 'ko-zh' else "한국어"
        source_lang = "한국어" if direction == 'ko-zh' else "중국어"

        gpt_prompt = f"""다음은 엑셀 파일에서 추출한 {source_lang} 텍스트들입니다. 각 텍스트를 {target_lang}로 번역해주세요. 영어는 그대로 유지해주세요.

번역할 텍스트 목록:
{chr(10).join(word_list)}

답변 형식:
각 줄마다 다음 형식으로 답변해주세요:
<셀주소, 원본텍스트> -> 번역된텍스트

예시:
<Sheet1!A1, 안녕하세요> -> 你好
<Sheet1!B2, 프로젝트 관리> -> 项目管理"""

        return jsonify({
            'success': True,
            'word_count': len(word_list),
            'word_list': word_list,
            'gpt_prompt': gpt_prompt,
            'job_id': unique_id,
            'backup_path': backup_path  # 백업 파일 경로 전달
        })

    except Exception as e:
        print(f"단어 추출 오류: {e}")
        return jsonify({'error': f'단어 추출 중 오류가 발생했습니다: {str(e)}'}), 500

@app.route('/process-gpt-translation', methods=['POST'])
def process_gpt_translation():
    """GPT 번역 결과를 받아서 엑셀 파일 번역 수행"""
    try:
        data = request.json
        gpt_response = data.get('gpt_response', '')
        job_id = data.get('job_id')  # extract-words에서 받은 job_id 사용
        direction = data.get('direction', 'ko-zh')
        preserve_english = data.get('preserve_english', True)
        add_new_sheet = data.get('add_new_sheet', True)

        if not gpt_response or not job_id:
            return jsonify({'error': 'GPT 응답과 파일 ID가 필요합니다.'}), 400

        # GPT 응답 파싱 - 줄바꿈 고려한 개선된 방식
        translation_map = {}

        print(f"GPT 응답 파싱 시작")
        print(f"GPT 응답 샘플: {gpt_response[:200]}...")

        # < 로 시작하는 항목들을 찾아서 파싱
        import re

        # <Sheet1!XXX, 원본텍스트> -> 번역텍스트 패턴 찾기
        # 줄바꿈이 있어도 처리할 수 있도록 DOTALL 플래그 사용
        pattern = r'<(Sheet\d*![A-Z]+\d+),\s*([^>]+)>\s*->\s*([^\n<]+)'

        matches = re.findall(pattern, gpt_response, re.DOTALL)

        print(f"정규표현식으로 찾은 매칭: {len(matches)}개")

        for i, (cell_address, original_text, translated_text) in enumerate(matches):
            # 줄바꿈과 공백 정리
            original_text = original_text.strip().replace('\n', ' ')
            translated_text = translated_text.strip()

            translation_map[cell_address] = translated_text
            print(f"  → 번역 매핑 {i+1}: {cell_address} = '{translated_text}'")

        # 정규표현식이 실패한 경우 기존 방식으로 백업 처리
        if len(translation_map) == 0:
            print("정규표현식 파싱 실패, 기존 방식으로 백업 처리")
            lines = gpt_response.strip().split('\n')

            current_entry = ""
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # 새로운 항목 시작 (<Sheet로 시작)
                if line.startswith('<Sheet'):
                    # 이전 항목 처리
                    if current_entry and ' -> ' in current_entry:
                        try:
                            parts = current_entry.split(' -> ', 1)
                            left_part = parts[0].strip()
                            translated_text = parts[1].strip()

                            if left_part.startswith('<') and '>' in left_part:
                                end_bracket = left_part.find('>')
                                if end_bracket != -1:
                                    bracket_content = left_part[1:end_bracket]
                                    if ', ' in bracket_content:
                                        cell_address, original_text = bracket_content.split(', ', 1)
                                        translation_map[cell_address] = translated_text
                                        print(f"  → 백업 매핑: {cell_address} = {translated_text}")
                        except Exception as e:
                            print(f"백업 파싱 오류: {current_entry} - {e}")

                    # 새 항목 시작
                    current_entry = line
                else:
                    # 기존 항목에 추가 (줄바꿈된 내용)
                    current_entry += " " + line

            # 마지막 항목 처리
            if current_entry and ' -> ' in current_entry:
                try:
                    parts = current_entry.split(' -> ', 1)
                    left_part = parts[0].strip()
                    translated_text = parts[1].strip()

                    if left_part.startswith('<') and '>' in left_part:
                        end_bracket = left_part.find('>')
                        if end_bracket != -1:
                            bracket_content = left_part[1:end_bracket]
                            if ', ' in bracket_content:
                                cell_address, original_text = bracket_content.split(', ', 1)
                                translation_map[cell_address] = translated_text
                                print(f"  → 백업 매핑 (마지막): {cell_address} = {translated_text}")
                except Exception as e:
                    print(f"백업 파싱 오류 (마지막): {current_entry} - {e}")

        print(f"파싱된 번역 맵: {len(translation_map)}개 항목")
        for addr, trans in list(translation_map.items())[:5]:  # 처음 5개만 출력
            print(f"  {addr}: {trans}")

        # tmp 디렉토리에서 백업 파일 찾기 및 번역 파일 생성
        import openpyxl
        import os
        import shutil
        import glob

        tmp_dir = os.path.join(os.getcwd(), 'tmp')

        # job_id로 백업 파일 찾기
        backup_pattern = os.path.join(tmp_dir, f"backup_{job_id}_*.xlsx")
        backup_files = glob.glob(backup_pattern)

        if not backup_files:
            return jsonify({'error': f'백업 파일을 찾을 수 없습니다. job_id: {job_id}'}), 400

        backup_path = backup_files[0]  # 첫 번째 매칭 파일 사용

        # 출력 파일 생성 (백업 파일을 완전히 복사)
        output_filename = f"translated_{job_id}.xlsx"
        output_path = os.path.join(os.getcwd(), output_filename)

        shutil.copy2(backup_path, output_path)
        print(f"백업 파일에서 번역 파일 생성: {backup_path} -> {output_path}")

        # 복사된 파일에서 번역 적용 - 간단하고 안정적인 방식
        workbook = openpyxl.load_workbook(output_path)
        applied_count = 0

        print(f"번역 적용 시작. 파일 시트: {workbook.sheetnames}")
        print(f"번역 맵에 있는 주소들: {list(translation_map.keys())[:10]}...")

        def find_translated_sheets(workbook):
            """복사된 번역 시트들을 찾는 함수"""
            translated_sheet_suffixes = ['_중문', '_translated', '_chinese', '_번역', '_zh', '_cn']
            translated_sheets = {}

            for sheet_name in workbook.sheetnames:
                for suffix in translated_sheet_suffixes:
                    if sheet_name.lower().endswith(suffix.lower()):
                        # 원본 시트명 추정
                        original_name = sheet_name[:-len(suffix)]
                        if original_name in workbook.sheetnames:
                            translated_sheets[original_name] = sheet_name
                            print(f"발견: '{original_name}' -> '{sheet_name}' (번역 시트)")
                        break

            return translated_sheets

        # 새 시트 생성 여부 확인
        if add_new_sheet:
            # 복사된 번역 시트 찾기
            translated_sheets = find_translated_sheets(workbook)

            if translated_sheets:
                print(f"복사된 번역 시트 발견: {len(translated_sheets)}개")
                print("번역 적용 시작 - 복사된 시트에만 번역 적용하여 원본 보존")

                for original_sheet, translated_sheet in translated_sheets.items():
                    print(f"처리 중: '{original_sheet}' -> '{translated_sheet}'")

                    # 해당 원본 시트의 번역 데이터만 필터링
                    sheet_translations = {addr: trans for addr, trans in translation_map.items()
                                        if addr.startswith(f"{original_sheet}!")}

                    if not sheet_translations:
                        print(f"  시트 '{original_sheet}' 번역 대상 없음")
                        continue

                    # 번역 시트에 적용
                    target_sheet = workbook[translated_sheet]
                    sheet_applied_count = 0

                    for row in target_sheet.iter_rows():
                        for cell in row:
                            # 원본 시트 기준으로 셀 주소 생성
                            cell_address = f"{original_sheet}!{cell.coordinate}"
                            if cell_address in translation_map:
                                old_value = cell.value
                                cell.value = translation_map[cell_address]
                                applied_count += 1
                                sheet_applied_count += 1
                                print(f"  번역 적용: {cell.coordinate} '{old_value}' -> '{translation_map[cell_address]}'")

                    print(f"  '{translated_sheet}' 완료: {sheet_applied_count}개 셀 번역")

                print(f"총 {applied_count}개 셀이 번역되었습니다.")

            else:
                print("복사된 번역 시트를 찾을 수 없습니다.")
                print("💡 사용법: 원본 시트를 복사하고 이름 끝에 '_중문', '_translated', '_chinese' 등을 추가하세요.")
                print("예시: 'Sheet1' -> 'Sheet1_중문' 또는 'Sheet1_translated'")
                print("원본 시트에 직접 번역을 적용합니다.")

                # 원본 시트에 직접 적용 (기존 로직)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_translations = {addr: trans for addr, trans in translation_map.items() if addr.startswith(f"{sheet_name}!")}

                    if not sheet_translations:
                        continue

                    print(f"시트 '{sheet_name}' 처리 중... ({len(sheet_translations)}개 번역 대상)")

                    for row in sheet.iter_rows():
                        for cell in row:
                            cell_address = f"{sheet_name}!{cell.coordinate}"
                            if cell_address in translation_map:
                                old_value = cell.value
                                cell.value = translation_map[cell_address]
                                applied_count += 1
                                print(f"번역 적용 {applied_count}: {sheet_name}!{cell.coordinate} '{old_value}' -> '{translation_map[cell_address]}'")

        else:
            # 원본 시트에 직접 번역 적용
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                print(f"시트 '{sheet_name}' 직접 번역 중...")

                # 번역 적용
                for row in sheet.iter_rows():
                    for cell in row:
                        cell_address = f"{sheet_name}!{cell.coordinate}"
                        if cell_address in translation_map:
                            old_value = cell.value
                            cell.value = translation_map[cell_address]
                            applied_count += 1
                            print(f"번역 적용 {applied_count}: {cell_address} '{old_value}' -> '{translation_map[cell_address]}'")

        print(f"총 {applied_count}개 셀에 번역 적용됨")

        # 수정된 파일 저장
        workbook.save(output_path)
        print(f"번역된 파일 저장 완료: {output_path}")

        # 작업 정보 저장 (job_id는 이미 정의됨)
        translation_jobs[job_id] = {
            'status': 'completed',
            'progress': 100,
            'message': 'GPT 번역 완료',
            'output_filename': output_filename,
            'original_filename': 'translated_file.xlsx',
            'result_path': output_path
        }

        return jsonify({
            'success': True,
            'job_id': job_id,
            'message': 'GPT 번역이 완료되었습니다.',
            'translations_applied': len(translation_map)
        })

    except Exception as e:
        print(f"GPT 번역 처리 오류: {e}")
        return jsonify({'error': f'GPT 번역 처리 중 오류가 발생했습니다: {str(e)}'}), 500

@app.route('/translate-excel', methods=['POST'])
def translate_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'error': '파일이 업로드되지 않았습니다.'}), 400

        file = request.files['file']
        direction = request.form.get('direction', 'ko-zh')
        preserve_english = request.form.get('preserve_english', 'true').lower() == 'true'
        add_new_sheet = request.form.get('add_new_sheet', 'true').lower() == 'true'

        # 번역 제외 설정 파싱
        exclude_sheets_str = request.form.get('exclude_sheets', '')
        exclude_cells_str = request.form.get('exclude_cells', '')
        exclude_patterns_str = request.form.get('exclude_patterns', '')

        exclude_sheets = [s.strip() for s in exclude_sheets_str.split(',') if s.strip()] if exclude_sheets_str else None
        exclude_cells = [c.strip() for c in exclude_cells_str.split(',') if c.strip()] if exclude_cells_str else None
        exclude_patterns = [p.strip() for p in exclude_patterns_str.split(',') if p.strip()] if exclude_patterns_str else None

        if exclude_cells:
            print(f"제외할 셀: {len(exclude_cells)}개")
        if exclude_sheets:
            print(f"제외할 시트: {exclude_sheets}")
        if exclude_patterns:
            print(f"제외할 패턴: {exclude_patterns}")

        if file.filename == '':
            return jsonify({'error': '파일이 선택되지 않았습니다.'}), 400

        # 파일 저장
        import os

        unique_id = str(uuid.uuid4())[:8]
        input_filename = f"input_{unique_id}_{file.filename}"
        output_filename = f"translated_{unique_id}_{file.filename}"

        input_path = os.path.join(os.getcwd(), input_filename)
        output_path = os.path.join(os.getcwd(), output_filename)

        file.save(input_path)

        # 번역 작업 정보 저장
        job_id = unique_id
        translation_jobs[job_id] = {
            'status': 'queued',
            'progress': 0,
            'message': '번역 준비 중...',
            'output_filename': output_filename,
            'original_filename': file.filename
        }

        # 백그라운드에서 번역 실행
        thread = threading.Thread(
            target=run_translation,
            args=(job_id, input_path, output_path, direction, preserve_english, add_new_sheet, exclude_sheets, exclude_cells, exclude_patterns)
        )
        thread.daemon = True
        thread.start()

        return jsonify({
            'success': True,
            'job_id': job_id,
            'message': '번역이 시작되었습니다.'
        })

    except Exception as e:
        print(f"엑셀 번역 오류: {e}")
        return jsonify({'error': f'번역 중 오류가 발생했습니다: {str(e)}'}), 500

@app.route('/translation-status/<job_id>', methods=['GET'])
def get_translation_status(job_id):
    """번역 상태 조회"""
    if job_id not in translation_jobs:
        return jsonify({'error': '작업을 찾을 수 없습니다.'}), 404
    
    job = translation_jobs[job_id]
    
    if job['status'] == 'completed':
        return jsonify({
            'status': job['status'],
            'progress': job['progress'],
            'message': job['message'],
            'download_filename': job['output_filename'],
            'file_id': job_id
        })
    elif job['status'] == 'error':
        return jsonify({
            'status': job['status'],
            'error': job.get('error', '알 수 없는 오류')
        })
    else:
        return jsonify({
            'status': job['status'],
            'progress': job['progress'],
            'message': job['message']
        })

@app.route('/download/<file_id>/<filename>', methods=['GET'])
def download_file(file_id, filename):
    try:
        from flask import send_file
        import os
        
        file_path = os.path.join(os.getcwd(), filename)
        
        if not os.path.exists(file_path):
            return jsonify({'error': '파일을 찾을 수 없습니다.'}), 404
        
        def remove_file():
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except:
                pass
        
        response = send_file(file_path, as_attachment=True, download_name=filename.replace(f'translated_{file_id}_', ''))
        
        # 파일 전송 후 정리 (1초 후)
        import threading
        timer = threading.Timer(1.0, remove_file)
        timer.start()
        
        return response
        
    except Exception as e:
        print(f"파일 다운로드 오류: {e}")
        return jsonify({'error': f'파일 다운로드 중 오류가 발생했습니다: {str(e)}'}), 500

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'message': '번역 서버가 정상 작동 중입니다.'})

if __name__ == '__main__':
    print("번역 서버 시작 중...")
    print("URL: http://localhost:5001")
    print("중지하려면 Ctrl+C를 누르세요.")
    app.run(host='0.0.0.0', port=5001, debug=True)